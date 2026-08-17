#!/usr/bin/env python3
"""Generate sample data files so the pipeline can be tested end-to-end
before real Capital IQ / faculty data is available.

Creates:
    data/companies_raw.xlsx      (Capital IQ screening layout, 12 companies)
    data/Faculty Database.xlsx   (2 faculty, Flag = N)

Safe to delete once real data is in place.
"""
import openpyxl

from nexus.settings import Settings

cfg = Settings()

# ---------------- companies_raw.xlsx (Capital IQ layout) ----------------
COMPANIES = [
    # name, employees, revenue, city/state, description, website
    ("Nutmeg Catalysts LLC", "45", "38.5", "100 Science Park Rd \nStorrs, Connecticut 06269",
     "Develops zeolite-based heterogeneous catalysts for fuel upgrading and petrochemical processes.",
     "www.nutmegcatalysts.example"),
    ("Bay State Pyrolysis Inc.", "120", "55.0", "12 Energy Way \nWorcester, Massachusetts 01605",
     "Waste-to-energy company operating thermochemical conversion and pyrolysis units for municipal waste.",
     "www.bspyro.example"),
    ("Ocean State Materials Corp.", "300", "210.4", "8 Harbor Blvd \nProvidence, Rhode Island 02903",
     "Advanced materials manufacturer specializing in porous ceramics and adsorbents.",
     "www.osmaterials.example"),
    ("Hartford BioPharma Co.", "5,200", "980.0", "1 Insurance Plaza \nHartford, Connecticut 06103",
     "Clinical-stage biopharmaceutical company focused on oncology therapeutics.",
     "www.hartfordbio.example"),
    ("Quiet Corner Insurance Group", "850", "450.0", "77 Main St \nPutnam, Connecticut 06260",
     "Regional property and casualty insurance provider.",
     "www.qcinsurance.example"),
    ("Springfield Robotics Ltd.", "60", "12.3", "300 Armory St \nSpringfield, Massachusetts 01105",
     "Designs autonomous mobile robots for warehouse logistics.",
     "www.sprobotics.example"),
    ("Thames River Fuels", "25", "-", "5 Dock Rd \nNew London, Connecticut 06320",
     "Produces renewable diesel and biofuels from waste oils using catalytic hydroprocessing.",
     "www.trfuels.example"),
    ("Berkshire Dental Supplies", "15", "4.2", "9 Elm St \nPittsfield, Massachusetts 01201",
     "Distributor of dental equipment and consumables.",
     "-"),
    ("Mystic Energy Storage", "80", "33.7", "200 Grid Ave \nMystic, Connecticut 06355",
     "Develops grid-scale battery storage and energy management systems.",
     "www.mysticenergy.example"),
    ("Providence Polymer Works", "220", "95.8", "41 Industrial Dr \nProvidence, Rhode Island 02905",
     "Manufactures specialty polymers and catalytic membrane materials for chemical processing.",
     "www.ppolymer.example"),
    ("Cape Cod Software LLC", "35", "8.9", "2 Beach Rd \nBarnstable, Massachusetts 02630",
     "SaaS platform for restaurant inventory management.",
     "www.ccsoft.example"),
    ("Storrs AgriScience Inc.", "-", "-", "45 Farm Rd \nStorrs, Connecticut 06268",
     "Agricultural biotechnology startup working on soil microbiome enhancement.",
     "-"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Screening"
ws["A6"] = "Capital IQ Company Screening Report > SAMPLE DATA (generated for testing)"
headers = {"A": "Company Name", "C": "Number of Employees - Global (Latest)",
           "D": "Revenue - Compustat [LTM] ($USDmm, Historical rate)",
           "G": "Offices", "I": "Business Description", "J": "Website"}
for col, text in headers.items():
    ws[f"{col}8"] = text
for i, (name, emp, rev, addr, desc, web) in enumerate(COMPANIES):
    r = 9 + i
    ws[f"A{r}"] = name
    ws[f"C{r}"] = emp
    ws[f"D{r}"] = rev
    ws[f"G{r}"] = f"Headquarters\n{addr} \nUnited States\nMain Phone: 555-0100"
    ws[f"I{r}"] = desc
    ws[f"J{r}"] = web
cfg.companies_raw.parent.mkdir(parents=True, exist_ok=True)
wb.save(cfg.companies_raw)
print(f"Sample datasource written: {cfg.companies_raw} ({len(COMPANIES)} companies)")

# ---------------- Faculty Database.xlsx ----------------
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Faculty"
ws2.append(["Faculty", "School/College", "Department", "Research Description",
            "Industry Classification1", "Industry Classification2", "Flag"])
ws2.append(["Ioulia Valla", "College of Engineering", "Chemical and Biological Engineering",
            "Heterogeneous catalysis and zeolite-based materials, Thermochemical conversion "
            "and pyrolysis and waste-to-energy processes, Fuel upgrading and catalyst "
            "performance under realistic operating conditions",
            "Energy", "Materials", "N"])
ws2.append(["Sample Professor", "School of Business", "Marketing",
            "Consumer behavior analytics and retail technology adoption",
            "Software", "Retail", "N"])
wb2.save(cfg.faculty_database)
print(f"Sample faculty database written: {cfg.faculty_database} (2 faculty, Flag=N)")
