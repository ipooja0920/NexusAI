"""All Excel reading/writing: Capital IQ export, enriched companies,
Faculty Database (including flag updates), and per-faculty output books."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


# ======================================================================
# Capital IQ raw export
# ======================================================================

def read_capitaliq(path: Path, cfg: dict) -> pd.DataFrame:
    """Read the raw Capital IQ screening export into a tidy DataFrame."""
    if not path.exists():
        raise FileNotFoundError(
            f"Company datasource not found at {path}. Place your Capital IQ "
            "export there (or update paths.companies_raw in config.yaml)."
        )
    usecols = ",".join([
        cfg["col_company"], cfg["col_employees"], cfg["col_revenue"],
        cfg["col_address"], cfg["col_description"], cfg["col_website"],
    ])
    df = pd.read_excel(
        path, header=None, skiprows=cfg["data_start_row"] - 1,
        usecols=usecols, engine="openpyxl", dtype=str,
    )
    df.columns = ["company", "employees", "revenue", "address", "description", "website"]

    def _clean(v) -> Optional[str]:
        """None / NaN / blank -> None; everything else -> stripped string."""
        if v is None:
            return None
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            return None
        return s

    rows = []
    for tup in df.itertuples(index=False, name=None):
        company = _clean(tup[0])
        if company is None:
            break  # stop at first blank company row (footer follows)
        rows.append({
            "company": company,
            "employees": _clean(tup[1]),
            "revenue": _clean(tup[2]),
            "address": _clean(tup[3]),
            "description": _clean(tup[4]),
            "website": _clean(tup[5]),
        })
    return pd.DataFrame(rows)


# ======================================================================
# Enriched companies file
# ======================================================================

ENRICHED_COLUMNS = [
    "company", "website", "description", "address_cleaned",
    "latitude", "longitude", "geocode_provider", "distance_miles",
    "employees", "revenue_usdmm",
    "distance_score", "employee_score", "revenue_score",
    "embedding_key",
]


def write_enriched(path: Path, records: List[dict]) -> None:
    df = pd.DataFrame(records, columns=ENRICHED_COLUMNS)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Companies")
        ws = writer.sheets["Companies"]
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.freeze_panes = "A2"


def read_enriched(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(
            f"Enriched company file not found at {path}. "
            "Run 'python enrich_companies.py' first."
        )
    return pd.read_excel(path, sheet_name="Companies", engine="openpyxl")


# ======================================================================
# Faculty Database
# ======================================================================

def read_faculty(path: Path, cfg: dict) -> List[dict]:
    if not path.exists():
        raise FileNotFoundError(
            f"Faculty Database not found at {path}. "
            "Place it there or update paths.faculty_database in config.yaml."
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    def col(key: str) -> int:
        return column_index_from_string(cfg[key])

    faculty = []
    for row in range(cfg["data_start_row"], ws.max_row + 1):
        name = ws.cell(row=row, column=col("col_name")).value
        if name is None or str(name).strip() == "":
            continue
        faculty.append({
            "row": row,
            "name": str(name).strip(),
            "school": str(ws.cell(row=row, column=col("col_school")).value or "").strip(),
            "department": str(ws.cell(row=row, column=col("col_department")).value or "").strip(),
            "research": str(ws.cell(row=row, column=col("col_research")).value or "").strip(),
            "class1": str(ws.cell(row=row, column=col("col_class1")).value or "").strip(),
            "class2": str(ws.cell(row=row, column=col("col_class2")).value or "").strip(),
            "flag": str(ws.cell(row=row, column=col("col_flag")).value or "").strip().upper(),
        })
    wb.close()
    return faculty


def set_faculty_flag(path: Path, cfg: dict, row: int, value: str = "Y") -> None:
    """Flip one faculty's flag and save. Retries advice on PermissionError
    (file open in Excel) is raised as a clear message."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.cell(row=row, column=column_index_from_string(cfg["col_flag"])).value = value
    try:
        wb.save(path)
    except PermissionError as e:
        raise PermissionError(
            f"Could not save {path.name} - is it open in Excel? "
            "Close it and re-run; completed faculty are already flagged."
        ) from e
    finally:
        wb.close()


# ======================================================================
# Per-faculty output workbook
# ======================================================================

RESULT_HEADERS = [
    "Rank", "Company", "Website", "Distance (mi)", "Employees", "Revenue ($USDmm)",
    "Distance Score", "Employee Score", "Revenue Score",
    "Alignment Score (1-9)", "Alignment Reason",
    "Partnership Score (1-9)", "Partnership Reason",
    "Funding Score (1-9)", "Funding Reason",
    "Overall Score", "Company Description",
]

COL_WIDTHS = {
    "A": 6, "B": 32, "C": 26, "D": 12, "E": 11, "F": 16,
    "G": 13, "H": 13, "I": 13, "J": 12, "K": 45, "L": 12, "M": 45,
    "N": 12, "O": 45, "P": 12, "Q": 60,
}


ALL_HEADERS = [
    "Stage", "Rank", "Company", "Website", "Distance (mi)", "Employees",
    "Revenue ($USDmm)", "Distance Score", "Employee Score", "Revenue Score",
    "Alignment Score (1-9)", "Alignment Reason",
    "Partnership Score (1-9)", "Partnership Reason",
    "Funding Score (1-9)", "Funding Reason",
    "Overall Score", "Company Description",
]

ALL_COL_WIDTHS = {
    "A": 22, "B": 6, "C": 32, "D": 26, "E": 12, "F": 11, "G": 16,
    "H": 13, "I": 13, "J": 13, "K": 12, "L": 45, "M": 12, "N": 45,
    "O": 12, "P": 45, "Q": 12, "R": 60,
}


def write_faculty_output(path: Path, faculty: dict, results: List[dict],
                         run_info: dict, all_rows: Optional[List[dict]] = None) -> None:
    wb = openpyxl.Workbook()

    # ---- Results sheet ----
    ws = wb.active
    ws.title = "Top Matches"
    ws.append(RESULT_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r in results:
        ws.append([
            r["rank"], r["company"], r.get("website"), r.get("distance_miles"),
            r.get("employees"), r.get("revenue_usdmm"),
            round(r["distance_score"], 4), round(r["employee_score"], 4),
            round(r["revenue_score"], 4),
            r.get("alignment_score"), r.get("alignment_reason"),
            r.get("partnership_score"), r.get("partnership_reason"),
            r.get("funding_score"), r.get("funding_reason"),
            round(r["overall_score"], 4), r.get("description"),
        ])

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "C2"

    # ---- All Companies sheet (full funnel visibility, v1-style) ----
    if all_rows:
        wa = wb.create_sheet("All Companies")
        wa.append(ALL_HEADERS)
        for cell in wa[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in all_rows:
            def _r4(v):
                return round(v, 4) if isinstance(v, (int, float)) else v
            wa.append([
                r.get("stage"), r.get("rank"), r["company"], r.get("website"),
                r.get("distance_miles"), r.get("employees"), r.get("revenue_usdmm"),
                _r4(r.get("distance_score")), _r4(r.get("employee_score")),
                _r4(r.get("revenue_score")),
                r.get("alignment_score"), r.get("alignment_reason"),
                r.get("partnership_score"), r.get("partnership_reason"),
                r.get("funding_score"), r.get("funding_reason"),
                _r4(r.get("overall_score")), r.get("description"),
            ])
        for letter, width in ALL_COL_WIDTHS.items():
            wa.column_dimensions[letter].width = width
        for row in wa.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        wa.freeze_panes = "D2"

    # ---- Run Info sheet (traceability) ----
    info = wb.create_sheet("Run Info")
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 100
    rows = [
        ("Faculty", faculty["name"]),
        ("School/College", faculty.get("school")),
        ("Department", faculty.get("department")),
        ("Research profile used", run_info.get("research_profile")),
        ("Run date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Scoring model", run_info.get("scoring_model")),
        ("Embedding model", run_info.get("embedding_model")),
        ("Companies in datasource", run_info.get("n_companies")),
        ("Candidates after pre-filter", run_info.get("n_candidates")),
        ("Secondary scoring cutoff", run_info.get("secondary_cutoff")),
        ("Weights", run_info.get("weights")),
    ]
    for k, v in rows:
        info.append([k, str(v) if v is not None else ""])
    for cell in info["A"]:
        cell.font = Font(bold=True)
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
