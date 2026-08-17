#!/usr/bin/env python3
"""End-to-end smoke test with mocked OpenAI + geocoding.

Verifies the full pipeline without network access or an API key:
  1. Pre-seeds the geocode cache (so Stage 0 'geocodes' from cache)
  2. Fakes the OpenAI client (embeddings + chat completions)
  3. Runs Stage 0 and Stage 1 end-to-end
  4. Checks outputs: enriched file, per-faculty workbooks, flags flipped

Run:  python test_smoke.py
(Developer tool - not needed for normal use.)
"""
import hashlib
import json
import sys

import openai as _openai_module

from nexus.cache import Caches
from nexus.excel_io import read_capitaliq, read_faculty
from nexus.geo import clean_address_block, select_address_block
from nexus.settings import Settings


# ----------------------------------------------------------------------
# Fake OpenAI client
# ----------------------------------------------------------------------
class _Msg:
    def __init__(self, content): self.content = content

class _Choice:
    def __init__(self, content): self.message = _Msg(content)

class _ChatResp:
    def __init__(self, content): self.choices = [_Choice(content)]

class _EmbData:
    def __init__(self, vec): self.embedding = vec

class _EmbResp:
    def __init__(self, vec): self.data = [_EmbData(vec)]


def _fake_embedding(text: str) -> list:
    """Deterministic pseudo-embedding; catalysis-ish texts cluster together."""
    h = hashlib.sha256(text.lower().encode()).digest()
    base = [b / 255.0 for b in h[:16]]
    keywords = ["cataly", "zeolite", "pyrolysis", "fuel", "thermochemical",
                "polymer", "energy", "materials"]
    signal = sum(1.0 for k in keywords if k in text.lower())
    return [signal] + base  # first dim = topical signal -> similar texts align


class FakeOpenAI:
    def __init__(self, *a, **k):
        self.chat = self._Chat()
        self.embeddings = self._Embeddings()

    class _Embeddings:
        def create(self, model, input):
            return _EmbResp(_fake_embedding(input))

    class _Chat:
        def __init__(self): self.completions = self

        def create(self, model, messages):
            user = messages[-1]["content"].lower()
            system = messages[0]["content"].lower()
            # crude scoring: catalysis-related companies score high
            hot = ["cataly", "zeolite", "pyrolysis", "fuel", "thermochemical", "polymer"]
            score = 8 if any(k in user for k in hot) else 3
            if "partner" in system:
                kind = "partnership"
            elif "budget" in system:
                kind = "funding"
            else:
                kind = "alignment"
            company = "Test Co"
            for line in messages[-1]["content"].splitlines():
                if line.startswith("Company:"):
                    company = line.split(":", 1)[1].strip()
            return _ChatResp(json.dumps({
                "company": company, "score": score,
                "reason": f"Mock {kind} reason for testing."
            }))


def main() -> int:
    print("=== NexusAI smoke test (mocked AI + geocoding) ===\n")
    cfg = Settings()

    # ---- 1. Pre-seed geocode cache from the raw file's addresses ----
    df = read_capitaliq(cfg.companies_raw, cfg.capitaliq)
    caches = Caches(cfg.cache_dir)
    for i, row in df.iterrows():
        cleaned = clean_address_block(select_address_block(row["address"]))
        if cleaned and caches.get_geocode(cleaned) is None:
            # spread fake coords around Storrs (i*0.1 deg ~ 7mi steps)
            caches.set_geocode(cleaned, 41.8073 + i * 0.1, -72.2536, "mock")
    caches.save_all()
    print(f"Seeded geocode cache: {len(caches.geocode)} addresses")

    # ---- 2. Patch OpenAI globally ----
    _openai_module.OpenAI = FakeOpenAI
    import os
    os.environ.setdefault("OPENAI_API_KEY", "sk-mock-for-smoke-test")

    # ---- 3. Stage 0 ----
    import enrich_companies
    sys.argv = ["enrich_companies.py"]
    rc = enrich_companies.main()
    assert rc == 0, f"Stage 0 failed with rc={rc}"

    # ---- 4. Stage 1 ----
    import match_faculty
    sys.argv = ["match_faculty.py"]
    rc = match_faculty.main()
    assert rc == 0, f"Stage 1 failed with rc={rc}"

    # ---- 5. Verify ----
    outputs = sorted(cfg.output_dir.glob("*.xlsx"))
    assert outputs, "No output workbooks produced"
    faculty = read_faculty(cfg.faculty_database, cfg.faculty)
    flags = {f["name"]: f["flag"] for f in faculty}
    assert all(v == "Y" for v in flags.values()), f"Flags not flipped: {flags}"

    import openpyxl
    wb = openpyxl.load_workbook(outputs[-1])
    ws = wb["Top Matches"]
    n_rows = ws.max_row - 1
    top_company = ws.cell(row=2, column=2).value
    print(f"\n--- VERIFICATION ---")
    print(f"Output workbooks: {[p.name for p in outputs]}")
    print(f"Flags after run: {flags}")
    print(f"Rows in latest output: {n_rows}; top match: {top_company}")
    assert "Run Info" in wb.sheetnames, "Run Info sheet missing"
    print("\nSMOKE TEST PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
