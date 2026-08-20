"""Tests for reading and writing Excel files.

Covers the real-world messiness of Capital IQ exports (blank cells arriving
as NaN, footer rows after the data, missing files) and verifies the faculty
flag is written without damaging the rest of the sheet.
"""
import unittest
from pathlib import Path

import openpyxl

from nexus.excel_io import (read_capitaliq, read_enriched, read_faculty,
                            set_faculty_flag, write_enriched,
                            write_faculty_output)
from tests.harness import TempProject


class TestReadCapitalIQ(unittest.TestCase):
    def test_reads_all_rows(self):
        with TempProject() as p:
            cfg = p.settings()
            df = read_capitaliq(cfg.companies_raw, cfg.capitaliq)
            self.assertEqual(len(df), 8)
            self.assertEqual(df.iloc[0]["company"], "Catalyst Works LLC")

    def test_blank_cells_become_none_not_nan(self):
        """Regression: real Capital IQ files deliver blanks as float NaN,
        which crashed the parser with 'float has no attribute strip'."""
        companies = [("Sparse Co.", None, None, None, None, None)]
        with TempProject(companies=companies) as p:
            cfg = p.settings()
            df = read_capitaliq(cfg.companies_raw, cfg.capitaliq)
            self.assertEqual(len(df), 1)
            row = df.iloc[0]
            self.assertEqual(row["company"], "Sparse Co.")
            for field in ["employees", "revenue", "description", "website"]:
                self.assertIsNone(row[field], f"{field} should be None")

    def test_stops_at_first_blank_company(self):
        """Capital IQ appends footer text below the data; the reader must
        stop at the first empty company cell rather than ingest it."""
        with TempProject() as p:
            cfg = p.settings()
            # add a gap then a stray footer row
            wb = openpyxl.load_workbook(cfg.companies_raw)
            ws = wb.active
            ws["A18"] = None
            ws["A19"] = "Data provided by Capital IQ"
            wb.save(cfg.companies_raw)
            df = read_capitaliq(cfg.companies_raw, cfg.capitaliq)
            self.assertEqual(len(df), 8)
            self.assertNotIn("Data provided by Capital IQ", list(df["company"]))

    def test_missing_file_raises_clear_error(self):
        with TempProject() as p:
            cfg = p.settings()
            cfg.companies_raw.unlink()
            with self.assertRaises(FileNotFoundError) as ctx:
                read_capitaliq(cfg.companies_raw, cfg.capitaliq)
            self.assertIn("Capital IQ", str(ctx.exception))

    def test_dash_values_preserved_as_dash(self):
        """'-' is Capital IQ's 'no data' marker and must survive to the
        scoring layer, which knows how to interpret it."""
        with TempProject() as p:
            cfg = p.settings()
            df = read_capitaliq(cfg.companies_raw, cfg.capitaliq)
            unknown = df[df["company"] == "Unknown Holdings Inc."].iloc[0]
            self.assertEqual(unknown["employees"], "-")
            self.assertEqual(unknown["description"], "-")


class TestFacultyDatabase(unittest.TestCase):
    def test_reads_faculty_fields(self):
        with TempProject() as p:
            cfg = p.settings()
            fac = read_faculty(cfg.faculty_database, cfg.faculty)
            self.assertEqual(len(fac), 1)
            f = fac[0]
            self.assertEqual(f["name"], "Ioulia Valla")
            self.assertEqual(f["department"], "Chemical and Biological Engineering")
            self.assertEqual(f["class1"], "Energy")
            self.assertEqual(f["flag"], "N")
            self.assertEqual(f["row"], 2)

    def test_flag_is_uppercased(self):
        faculty = [("Test Prof", "School", "Dept", "Research", "C1", "C2", "n")]
        with TempProject(faculty=faculty) as p:
            cfg = p.settings()
            self.assertEqual(read_faculty(cfg.faculty_database, cfg.faculty)[0]["flag"], "N")

    def test_blank_rows_skipped(self):
        with TempProject() as p:
            cfg = p.settings()
            wb = openpyxl.load_workbook(cfg.faculty_database)
            ws = wb.active
            ws["A5"] = None
            ws["A6"] = "Later Professor"
            ws["G6"] = "N"
            wb.save(cfg.faculty_database)
            fac = read_faculty(cfg.faculty_database, cfg.faculty)
            self.assertEqual([f["name"] for f in fac], ["Ioulia Valla", "Later Professor"])
            self.assertEqual(fac[1]["row"], 6)   # row number must stay accurate

    def test_set_flag_writes_and_preserves_other_cells(self):
        with TempProject() as p:
            cfg = p.settings()
            set_faculty_flag(cfg.faculty_database, cfg.faculty, 2, "Y")
            fac = read_faculty(cfg.faculty_database, cfg.faculty)[0]
            self.assertEqual(fac["flag"], "Y")
            # everything else untouched
            self.assertEqual(fac["name"], "Ioulia Valla")
            self.assertEqual(fac["school"], "College of Engineering")
            self.assertIn("catalysis", fac["research"])

    def test_missing_faculty_file_raises_clear_error(self):
        with TempProject() as p:
            cfg = p.settings()
            cfg.faculty_database.unlink()
            with self.assertRaises(FileNotFoundError) as ctx:
                read_faculty(cfg.faculty_database, cfg.faculty)
            self.assertIn("Faculty Database", str(ctx.exception))


class TestEnrichedRoundTrip(unittest.TestCase):
    def test_write_then_read(self):
        with TempProject() as p:
            cfg = p.settings()
            records = [{
                "company": "Test Co", "website": "www.t.example", "description": "d",
                "address_cleaned": "1 Main St, USA", "latitude": 41.8, "longitude": -72.2,
                "geocode_provider": "test", "distance_miles": 12.5,
                "employees": 50, "revenue_usdmm": 40.0,
                "distance_score": 0.97, "employee_score": 0.05, "revenue_score": 1.0,
                "embedding_key": "abc123",
            }]
            write_enriched(cfg.companies_enriched, records)
            df = read_enriched(cfg.companies_enriched)
            self.assertEqual(len(df), 1)
            self.assertEqual(df.iloc[0]["company"], "Test Co")
            self.assertAlmostEqual(df.iloc[0]["distance_miles"], 12.5)

    def test_missing_enriched_file_tells_you_to_run_stage_0(self):
        with TempProject() as p:
            cfg = p.settings()
            with self.assertRaises(FileNotFoundError) as ctx:
                read_enriched(cfg.companies_enriched)
            self.assertIn("enrich_companies.py", str(ctx.exception))


class TestUniquePath(unittest.TestCase):
    """Regression: two runs for the same faculty in the same second used to
    produce the same filename, silently overwriting the earlier output."""

    def test_returns_path_unchanged_when_free(self):
        from nexus.excel_io import unique_path
        with TempProject() as p:
            target = p.dir / "report.xlsx"
            self.assertEqual(unique_path(target), target)

    def test_adds_suffix_when_taken(self):
        from nexus.excel_io import unique_path
        with TempProject() as p:
            target = p.dir / "report.xlsx"
            target.write_text("x")
            self.assertEqual(unique_path(target).name, "report (2).xlsx")

    def test_increments_past_multiple_collisions(self):
        from nexus.excel_io import unique_path
        with TempProject() as p:
            (p.dir / "report.xlsx").write_text("x")
            (p.dir / "report (2).xlsx").write_text("x")
            self.assertEqual(unique_path(p.dir / "report.xlsx").name, "report (3).xlsx")


class TestOutputWorkbook(unittest.TestCase):
    def _sample(self):
        faculty = {"name": "Ioulia Valla", "school": "CoE", "department": "CBE"}
        results = [{
            "rank": 1, "company": "Catalyst Works LLC", "website": "w",
            "distance_miles": 5.0, "employees": 45, "revenue_usdmm": 38.5,
            "distance_score": 0.98, "employee_score": 0.045, "revenue_score": 1.0,
            "alignment_score": "8", "alignment_reason": "strong",
            "partnership_score": "7", "partnership_reason": "likely",
            "funding_score": "5", "funding_reason": "moderate",
            "overall_score": 0.7123, "description": "desc",
        }]
        all_rows = results[0].copy()
        all_rows["stage"] = "Top 20"
        run_info = {"research_profile": "catalysis", "scoring_model": "m",
                    "embedding_model": "e", "n_companies": 8, "n_candidates": 8,
                    "secondary_cutoff": 40, "weights": {"distance": 0.25}}
        return faculty, results, [all_rows], run_info

    def test_creates_three_sheets(self):
        faculty, results, all_rows, run_info = self._sample()
        with TempProject() as p:
            out = p.dir / "out.xlsx"
            write_faculty_output(out, faculty, results, run_info, all_rows=all_rows)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb.sheetnames, ["Top Matches", "All Companies", "Run Info"])

    def test_omits_all_companies_sheet_when_not_supplied(self):
        faculty, results, _, run_info = self._sample()
        with TempProject() as p:
            out = p.dir / "out.xlsx"
            write_faculty_output(out, faculty, results, run_info)
            wb = openpyxl.load_workbook(out)
            self.assertEqual(wb.sheetnames, ["Top Matches", "Run Info"])

    def test_top_matches_values(self):
        faculty, results, all_rows, run_info = self._sample()
        with TempProject() as p:
            out = p.dir / "out.xlsx"
            write_faculty_output(out, faculty, results, run_info, all_rows=all_rows)
            ws = openpyxl.load_workbook(out)["Top Matches"]
            self.assertEqual(ws.cell(1, 1).value, "Rank")
            self.assertEqual(ws.cell(2, 1).value, 1)
            self.assertEqual(ws.cell(2, 2).value, "Catalyst Works LLC")
            self.assertEqual(ws.cell(2, 16).value, 0.7123)

    def test_run_info_records_profile_and_weights(self):
        faculty, results, all_rows, run_info = self._sample()
        with TempProject() as p:
            out = p.dir / "out.xlsx"
            write_faculty_output(out, faculty, results, run_info, all_rows=all_rows)
            ws = openpyxl.load_workbook(out)["Run Info"]
            labels = {ws.cell(r, 1).value: ws.cell(r, 2).value
                      for r in range(1, ws.max_row + 1)}
            self.assertEqual(labels["Faculty"], "Ioulia Valla")
            self.assertEqual(labels["Research profile used"], "catalysis")
            self.assertIn("distance", labels["Weights"])


if __name__ == "__main__":
    unittest.main()
