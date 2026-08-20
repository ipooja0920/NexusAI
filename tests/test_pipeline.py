"""End-to-end pipeline tests, run against a throwaway project.

The AI client and geocoder are stubbed, so these tests need no API key and
make no network calls. They verify the behaviours that matter operationally:
flags flip, interrupted runs resume, caches are reused, one bad faculty
doesn't sink the rest, and --dry-run really spends nothing.
"""
import unittest

import openpyxl

from tests.harness import FakeOpenAI, TempProject, install_fake_openai


class PipelineTestCase(unittest.TestCase):
    def setUp(self):
        install_fake_openai()

    @staticmethod
    def _prepared(**kwargs):
        """A temp project with geocodes already cached (no network needed)."""
        p = TempProject(**kwargs)
        p.__enter__()
        p.seed_geocodes()
        return p


class TestFullRun(PipelineTestCase):
    def test_enrich_then_match_produces_output(self):
        p = self._prepared()
        try:
            self.assertEqual(p.run_enrich(), 0)
            self.assertTrue(p.settings().companies_enriched.exists())

            self.assertEqual(p.run_match(), 0)
            outputs = p.outputs()
            self.assertEqual(len(outputs), 1)
            self.assertIn("Ioulia Valla", outputs[0].name)

            wb = openpyxl.load_workbook(outputs[0])
            self.assertIn("Top Matches", wb.sheetnames)
            self.assertIn("All Companies", wb.sheetnames)
            self.assertIn("Run Info", wb.sheetnames)
        finally:
            p.__exit__()

    def test_flag_flips_to_Y(self):
        p = self._prepared()
        try:
            p.run_enrich()
            self.assertEqual(p.faculty_flags()["Ioulia Valla"], "N")
            p.run_match()
            self.assertEqual(p.faculty_flags()["Ioulia Valla"], "Y")
        finally:
            p.__exit__()

    def test_relevant_companies_outrank_irrelevant_ones(self):
        """Sanity check on the whole scoring chain: a catalysis company
        should beat a dental supplier for a catalysis professor."""
        p = self._prepared()
        try:
            p.run_enrich()
            p.run_match()
            ws = openpyxl.load_workbook(p.outputs()[0])["Top Matches"]
            ranked = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
            self.assertIn("Catalyst Works LLC", ranked)
            pos = {name: i for i, name in enumerate(ranked)}
            for good in ["Catalyst Works LLC", "Pyro Energy Inc."]:
                for bad in ["Dental Supplies Co.", "Quiet Insurance Group"]:
                    if good in pos and bad in pos:
                        self.assertLess(pos[good], pos[bad],
                                        f"{good} should outrank {bad}")
        finally:
            p.__exit__()

    def test_all_companies_sheet_covers_every_company(self):
        p = self._prepared()
        try:
            p.run_enrich()
            p.run_match()
            ws = openpyxl.load_workbook(p.outputs()[0])["All Companies"]
            self.assertEqual(ws.max_row - 1, 8)          # all 8 fixtures present
            stages = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
            self.assertTrue(stages <= {"Top 20", "Secondary scored (top 40)",
                                       "Alignment scored", "Filtered out (pre-filter)"})
        finally:
            p.__exit__()


class TestResumeAndSelection(PipelineTestCase):
    def test_second_run_skips_completed_faculty(self):
        p = self._prepared()
        try:
            p.run_enrich()
            p.run_match()
            self.assertEqual(len(p.outputs()), 1)
            # Re-run: nothing left with Flag = N
            self.assertEqual(p.run_match(), 0)
            self.assertEqual(len(p.outputs()), 1, "should not re-process a Y faculty")
        finally:
            p.__exit__()

    def test_faculty_flag_overrides_and_reruns(self):
        p = self._prepared()
        try:
            p.run_enrich()
            p.run_match()
            p.run_match(["--faculty", "Ioulia Valla"])
            self.assertEqual(len(p.outputs()), 2, "--faculty should ignore the Y flag")
        finally:
            p.__exit__()

    def test_unknown_faculty_name_errors_cleanly(self):
        p = self._prepared()
        try:
            p.run_enrich()
            rc = p.run_match(["--faculty", "Nobody At All"])
            self.assertEqual(rc, 1)
            self.assertEqual(len(p.outputs()), 0)
        finally:
            p.__exit__()

    def test_dry_run_spends_no_api_calls(self):
        p = self._prepared()
        try:
            p.run_enrich()
            FakeOpenAI.reset()
            rc = p.run_match(["--dry-run"])
            self.assertEqual(rc, 0)
            self.assertEqual(FakeOpenAI.counts(), {}, "dry-run must call no APIs")
            self.assertEqual(len(p.outputs()), 0)
            self.assertEqual(p.faculty_flags()["Ioulia Valla"], "N")
        finally:
            p.__exit__()

    def test_multiple_faculty_each_get_a_file(self):
        faculty = [
            ("Prof One", "CoE", "CBE", "catalysis and zeolite materials", "Energy", "", "N"),
            ("Prof Two", "Business", "Marketing", "consumer retail analytics", "Retail", "", "N"),
            ("Prof Three", "CoE", "CBE", "pyrolysis and fuels", "Energy", "", "Y"),
        ]
        p = self._prepared(faculty=faculty)
        try:
            p.run_enrich()
            p.run_match()
            names = [o.name for o in p.outputs()]
            self.assertEqual(len(names), 2, "only the two Flag=N faculty run")
            self.assertTrue(any("Prof One" in n for n in names))
            self.assertTrue(any("Prof Two" in n for n in names))
            flags = p.faculty_flags()
            self.assertEqual(flags["Prof One"], "Y")
            self.assertEqual(flags["Prof Two"], "Y")
        finally:
            p.__exit__()


class TestCaching(PipelineTestCase):
    def test_secondary_scores_reused_across_faculty(self):
        """Partnership/funding answers don't depend on the professor, so the
        second faculty should trigger far fewer of those calls."""
        faculty = [
            ("Prof One", "CoE", "CBE", "catalysis and zeolites", "Energy", "", "N"),
            ("Prof Two", "CoE", "CBE", "catalysis and fuels", "Energy", "", "N"),
        ]
        p = self._prepared(faculty=faculty)
        try:
            p.run_enrich()
            FakeOpenAI.reset()
            p.run_match()
            counts = FakeOpenAI.counts()
            # 8 companies x 2 faculty = 16 alignment calls (no caching there)
            self.assertEqual(counts.get("alignment", 0), 16)
            # secondary scored once per company only, not once per faculty
            self.assertEqual(counts.get("partnership", 0), 8)
            self.assertEqual(counts.get("funding", 0), 8)
        finally:
            p.__exit__()

    def test_re_running_enrich_uses_cache_and_makes_no_new_calls(self):
        p = self._prepared()
        try:
            p.run_enrich()
            FakeOpenAI.reset()
            p.run_enrich()
            self.assertEqual(FakeOpenAI.counts().get("embedding", 0), 0,
                             "second enrichment should be fully cached")
        finally:
            p.__exit__()

    def test_changed_description_re_embeds(self):
        """Cache is keyed by text hash, so edited descriptions refresh."""
        p = self._prepared()
        try:
            p.run_enrich()
            cfg = p.settings()
            wb = openpyxl.load_workbook(cfg.companies_raw)
            wb.active["I9"] = "Completely rewritten description about pyrolysis."
            wb.save(cfg.companies_raw)
            FakeOpenAI.reset()
            p.run_enrich()
            self.assertEqual(FakeOpenAI.counts().get("embedding", 0), 1,
                             "only the edited company should re-embed")
        finally:
            p.__exit__()


class TestPrefilterActivation(PipelineTestCase):
    def test_prefilter_inactive_when_fewer_companies_than_cutoff(self):
        p = self._prepared()          # 8 companies, cutoff 300
        try:
            p.run_enrich()
            FakeOpenAI.reset()
            p.run_match()
            counts = FakeOpenAI.counts()
            self.assertEqual(counts.get("alignment", 0), 8, "all companies scored")
            self.assertEqual(counts.get("embedding", 0), 0,
                             "faculty profile should not be embedded when filter is off")
        finally:
            p.__exit__()

    def test_prefilter_active_when_cutoff_is_small(self):
        p = self._prepared(config_overrides={"prefilter": {"top_n_candidates": 3}})
        try:
            p.run_enrich()
            FakeOpenAI.reset()
            p.run_match()
            counts = FakeOpenAI.counts()
            self.assertEqual(counts.get("alignment", 0), 3, "only candidates scored")
            self.assertEqual(counts.get("embedding", 0), 1, "faculty profile embedded once")

            ws = openpyxl.load_workbook(p.outputs()[0])["All Companies"]
            stages = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
            self.assertEqual(stages.count("Filtered out (pre-filter)"), 5)
        finally:
            p.__exit__()

    def test_prefilter_keeps_the_relevant_companies(self):
        p = self._prepared(config_overrides={"prefilter": {"top_n_candidates": 3}})
        try:
            p.run_enrich()
            p.run_match()
            ws = openpyxl.load_workbook(p.outputs()[0])["All Companies"]
            scored = {ws.cell(r, 3).value for r in range(2, ws.max_row + 1)
                      if ws.cell(r, 1).value != "Filtered out (pre-filter)"}
            self.assertIn("Catalyst Works LLC", scored)
            self.assertNotIn("Dental Supplies Co.", scored)
        finally:
            p.__exit__()


class TestRobustness(PipelineTestCase):
    def test_na_and_malformed_ai_responses_do_not_crash(self):
        """'Unknown Holdings' returns NA; 'Broken Data Corp' returns junk.
        Both must be handled without stopping the run."""
        p = self._prepared()
        try:
            p.run_enrich()
            self.assertEqual(p.run_match(), 0)
            ws = openpyxl.load_workbook(p.outputs()[0])["All Companies"]
            rows = {ws.cell(r, 3).value: ws.cell(r, 11).value
                    for r in range(2, ws.max_row + 1)}
            self.assertEqual(rows["Unknown Holdings Inc."], "NA")
            self.assertEqual(rows["Broken Data Corp."], "NA")
        finally:
            p.__exit__()

    def test_match_without_enrichment_fails_clearly(self):
        p = self._prepared()
        try:
            with self.assertRaises(FileNotFoundError) as ctx:
                p.run_match()
            self.assertIn("enrich_companies.py", str(ctx.exception))
        finally:
            p.__exit__()

    def test_company_with_no_geocode_still_scored(self):
        """A failed geocode must not remove a company from the run;
        it just loses the distance component."""
        companies = [("No Address Co.", "10", "40", None,
                      "Zeolite catalyst manufacturer.", "-")]
        p = TempProject(companies=companies)
        p.__enter__()
        try:
            p.run_enrich(["--skip-geocode"])
            p.run_match()
            ws = openpyxl.load_workbook(p.outputs()[0])["Top Matches"]
            self.assertEqual(ws.cell(2, 2).value, "No Address Co.")
            self.assertIsNone(ws.cell(2, 4).value)      # distance blank
            self.assertEqual(ws.cell(2, 7).value, 0)    # distance score 0
        finally:
            p.__exit__()

    def test_empty_faculty_list_exits_cleanly(self):
        faculty = [("Done Prof", "S", "D", "research", "C1", "C2", "Y")]
        p = self._prepared(faculty=faculty)
        try:
            p.run_enrich()
            self.assertEqual(p.run_match(), 0)
            self.assertEqual(len(p.outputs()), 0)
        finally:
            p.__exit__()


if __name__ == "__main__":
    unittest.main()
