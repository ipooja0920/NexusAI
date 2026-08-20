"""Test harness: builds a throwaway NexusAI project in a temp directory.

Every test runs against its own isolated copy — its own config.yaml, its own
data/ folder, its own caches. Nothing here can read or write the real
data/ folder, so running the tests can never corrupt your geocode cache,
overwrite outputs, or flip a faculty flag.

Also provides FakeOpenAI, a stand-in client so tests need no API key and
make no network calls.
"""
from __future__ import annotations

import hashlib
import json
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl
import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent


# ======================================================================
# Fake OpenAI client (no network, no key, deterministic)
# ======================================================================

class _Msg:
    def __init__(self, content): self.content = content


class _Choice:
    def __init__(self, content): self.message = _Msg(content)


class _ChatResp:
    def __init__(self, content): self.choices = [_Choice(content)]


class _EmbData:
    def __init__(self, vec): self.embedding = vec


class _EmbResp:
    def __init__(self, vec): self.data = [_EmbData(vec)]


def fake_embedding(text: str) -> list:
    """Deterministic pseudo-embedding. Dimension 0 carries a topical signal
    so that catalysis-ish texts genuinely rank near a catalysis profile."""
    keywords = ["cataly", "zeolite", "pyrolysis", "fuel", "thermochemical",
                "polymer", "nanomaterial", "energy", "materials"]
    signal = sum(1.0 for k in keywords if k in text.lower())
    h = hashlib.sha256(text.lower().encode()).digest()
    return [signal] + [b / 255.0 for b in h[:16]]


class FakeOpenAI:
    """Drop-in replacement for openai.OpenAI. Counts calls so tests can
    assert on API usage (e.g. that --dry-run spends nothing)."""

    call_log: list = []

    def __init__(self, *a, **k):
        self.chat = self._Chat()
        self.embeddings = self._Embeddings()

    @classmethod
    def reset(cls):
        cls.call_log = []

    @classmethod
    def counts(cls) -> dict:
        out = {}
        for kind in cls.call_log:
            out[kind] = out.get(kind, 0) + 1
        return out

    class _Embeddings:
        def create(self, model, input):
            FakeOpenAI.call_log.append("embedding")
            return _EmbResp(fake_embedding(input))

    class _Chat:
        def __init__(self): self.completions = self

        def create(self, model, messages):
            system = messages[0]["content"].lower()
            user = messages[-1]["content"]
            company = ""
            for line in user.splitlines():
                if line.startswith("Company:"):
                    company = line.split(":", 1)[1].strip()

            if "partner" in system:
                kind = "partnership"
            elif "budget" in system:
                kind = "funding"
            else:
                kind = "alignment"
            FakeOpenAI.call_log.append(kind)

            # Companies named "*Unknown*" return NA, to exercise that path
            if "unknown" in company.lower():
                return _ChatResp(json.dumps({
                    "company": company, "score": "NA",
                    "reason": "No alignment information available"}))
            # Deliberately malformed for "*Broken*", to exercise the parser
            if "broken" in company.lower():
                return _ChatResp("this is not json at all")

            hot = ["cataly", "zeolite", "pyrolysis", "fuel", "thermochemical",
                   "polymer", "nanomaterial"]
            score = 8 if any(k in user.lower() for k in hot) else 3
            return _ChatResp(json.dumps({
                "company": company, "score": score,
                "reason": f"Mock {kind} reason."}))


def install_fake_openai():
    """Patch openai.OpenAI globally and set a dummy key."""
    import os
    import openai
    openai.OpenAI = FakeOpenAI
    os.environ["OPENAI_API_KEY"] = "sk-test-not-a-real-key"
    FakeOpenAI.reset()


# ======================================================================
# Isolated project builder
# ======================================================================

DEFAULT_COMPANIES = [
    # name, employees, revenue, city/state, description, website
    ("Catalyst Works LLC", "45", "38.5", "1 Science Park\nStorrs, Connecticut 06269",
     "Develops zeolite catalysts for fuel upgrading and pyrolysis processes.",
     "www.catalystworks.example"),
    ("Pyro Energy Inc.", "120", "55.0", "12 Energy Way\nWorcester, Massachusetts 01605",
     "Thermochemical conversion and waste-to-energy pyrolysis systems.",
     "www.pyroenergy.example"),
    ("Polymer Materials Corp.", "300", "210.4", "8 Harbor Blvd\nProvidence, Rhode Island 02903",
     "Manufactures specialty polymers and catalytic membrane materials.",
     "www.polymat.example"),
    ("Quiet Insurance Group", "850", "450.0", "77 Main St\nPutnam, Connecticut 06260",
     "Regional property and casualty insurance provider.",
     "www.qig.example"),
    ("Dental Supplies Co.", "15", "4.2", "9 Elm St\nPittsfield, Massachusetts 01201",
     "Distributor of dental equipment and consumables.", "-"),
    ("Vague Chemical LLC", "25", "-", "82 Crenshaw Dr\nFlanders, New Jersey 07836",
     "Vague Chemical LLC is a diversified chemicals company. It is based in Flanders, New Jersey.",
     "www.vaguechem.example"),
    ("Unknown Holdings Inc.", "-", "-", "New York\nUnited States", "-", "-"),
    ("Broken Data Corp.", "50", "12.0", "5 Test Rd\nHartford, Connecticut 06103",
     "A company used to test malformed AI responses.", "-"),
]

DEFAULT_FACULTY = [
    ("Ioulia Valla", "College of Engineering", "Chemical and Biological Engineering",
     "Heterogeneous catalysis and zeolite-based materials, Thermochemical conversion "
     "and pyrolysis and waste-to-energy processes, Fuel upgrading",
     "Energy", "Materials", "N"),
]


class TempProject:
    """Creates an isolated project directory with config + data files.

    Usage:
        with TempProject() as p:
            p.settings()          # Settings object pointed at the temp dir
            p.run_enrich()
            p.run_match()
    """

    def __init__(self, companies=None, faculty=None, config_overrides=None):
        self.companies = companies if companies is not None else DEFAULT_COMPANIES
        self.faculty = faculty if faculty is not None else DEFAULT_FACULTY
        self.config_overrides = config_overrides or {}
        self.dir = None

    def __enter__(self):
        self.dir = Path(tempfile.mkdtemp(prefix="nexusai_test_"))
        (self.dir / "data" / "cache").mkdir(parents=True)
        (self.dir / "data" / "output").mkdir(parents=True)
        self._write_config()
        self._write_companies()
        self._write_faculty()
        return self

    def __exit__(self, *exc):
        if self.dir and self.dir.exists():
            shutil.rmtree(self.dir, ignore_errors=True)
        return False

    # ---------- file builders ----------
    def _write_config(self):
        with open(PROJECT_ROOT / "config.yaml", encoding="utf-8") as fh:
            cfg = yaml.safe_load(fh)
        # Point every path at the temp dir (absolute, so no ambiguity)
        cfg["paths"] = {
            "companies_raw": str(self.dir / "data" / "companies_raw.xlsx"),
            "companies_enriched": str(self.dir / "data" / "companies_enriched.xlsx"),
            "faculty_database": str(self.dir / "data" / "Faculty Database.xlsx"),
            "output_dir": str(self.dir / "data" / "output"),
            "cache_dir": str(self.dir / "data" / "cache"),
        }
        for section, values in self.config_overrides.items():
            cfg.setdefault(section, {}).update(values)
        with open(self.dir / "config.yaml", "w", encoding="utf-8") as fh:
            yaml.safe_dump(cfg, fh)

    def _write_companies(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Screening"
        ws["A6"] = "Capital IQ Company Screening Report > TEST FIXTURE"
        headers = {"A": "Company Name", "C": "Number of Employees - Global (Latest)",
                   "D": "Revenue ($USDmm)", "G": "Offices",
                   "I": "Business Description", "J": "Website"}
        for col, text in headers.items():
            ws[f"{col}8"] = text
        for i, (name, emp, rev, addr, desc, web) in enumerate(self.companies):
            r = 9 + i
            ws[f"A{r}"] = name
            ws[f"C{r}"] = emp
            ws[f"D{r}"] = rev
            ws[f"G{r}"] = f"Headquarters\n{addr}\nUnited States\nMain Phone: 555-0100"
            ws[f"I{r}"] = desc
            ws[f"J{r}"] = web
        wb.save(self.dir / "data" / "companies_raw.xlsx")

    def _write_faculty(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Faculty", "School/College", "Department", "Research Description",
                   "Industry Classification1", "Industry Classification2", "Flag"])
        for row in self.faculty:
            ws.append(list(row))
        wb.save(self.dir / "data" / "Faculty Database.xlsx")

    # ---------- accessors ----------
    def settings(self):
        from nexus.settings import Settings
        return Settings(self.dir / "config.yaml")

    def seed_geocodes(self):
        """Pre-fill the geocode cache so no network call is needed."""
        from nexus.cache import Caches
        from nexus.excel_io import read_capitaliq
        from nexus.geo import clean_address_block, select_address_block
        cfg = self.settings()
        caches = Caches(cfg.cache_dir)
        df = read_capitaliq(cfg.companies_raw, cfg.capitaliq)
        for i, row in df.iterrows():
            cleaned = clean_address_block(select_address_block(row["address"]))
            if cleaned:
                # spread test coords in a predictable arc around Storrs
                caches.set_geocode(cleaned, 41.8073 + i * 0.05, -72.2536, "test")
        caches.save_all()

    def outputs(self):
        return sorted((self.dir / "data" / "output").glob("*.xlsx"))

    def faculty_flags(self) -> dict:
        from nexus.excel_io import read_faculty
        cfg = self.settings()
        return {f["name"]: f["flag"] for f in read_faculty(cfg.faculty_database, cfg.faculty)}

    # ---------- runners ----------
    def _run(self, module_name: str, script: str, argv):
        """Point NEXUSAI_CONFIG at this temp project and run a CLI main()."""
        import importlib
        import os
        mod = importlib.import_module(module_name)
        old_env = os.environ.get("NEXUSAI_CONFIG")
        old_argv = sys.argv
        os.environ["NEXUSAI_CONFIG"] = str(self.dir / "config.yaml")
        sys.argv = [script] + (argv or [])
        try:
            return mod.main()
        finally:
            sys.argv = old_argv
            if old_env is None:
                os.environ.pop("NEXUSAI_CONFIG", None)
            else:
                os.environ["NEXUSAI_CONFIG"] = old_env

    def run_enrich(self, argv=None) -> int:
        """Run Stage 0 against this temp project."""
        return self._run("enrich_companies", "enrich_companies.py", argv)

    def run_match(self, argv=None) -> int:
        """Run Stage 1 against this temp project."""
        return self._run("match_faculty", "match_faculty.py", argv)
